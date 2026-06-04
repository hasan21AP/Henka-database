import os
from urllib.parse import quote
from openpyxl import load_workbook

# Paths
songs_folder = r"G:\Visual_Studio_Work_Place\Python Projects\downloads"
excel_file = r"G:\Visual_Studio_Work_Place\Python Projects\anime_songs_questions.xlsx"

# Firebase base URL
firebase_base = (
    "https://firebasestorage.googleapis.com/v0/b/"
    "henka-game.firebasestorage.app/o/anime_songs%2F"
)

# Firebase suffix
firebase_suffix = (
    "?alt=media&token=daf18a20-80e2-4f5a-b526-9842c1c8ee3a"
)

# Load workbook
wb = load_workbook(excel_file)
ws = wb.active

# Get existing answers to avoid duplicates
existing_answers = set()

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1]:
        existing_answers.add(str(row[1]).strip())

# Find next empty row
next_row = ws.max_row + 1

# Read mp3 files
files = [f for f in os.listdir(songs_folder) if f.lower().endswith(".mp3")]

added_count = 0

for file_name in files:

    # Remove extension
    answer = os.path.splitext(file_name)[0].strip()

    # Skip duplicates
    if answer in existing_answers:
        print(f"Skipped duplicate: {answer}")
        continue

    # Encode filename for URL
    encoded_name = quote(file_name)

    # Build Firebase URL
    firebase_url = f"{firebase_base}{encoded_name}{firebase_suffix}"

    # Write data
    ws.cell(next_row, 1).value = "في أي أنمي هذه الأغنية؟"
    ws.cell(next_row, 2).value = answer
    ws.cell(next_row, 3).value = firebase_url
    ws.cell(next_row, 4).value = 100

    print(f"Added: {answer}")

    next_row += 1
    added_count += 1

# Save file
wb.save(excel_file)

print(f"\nDone! Added {added_count} songs.")