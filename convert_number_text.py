from openpyxl import load_workbook

# File path
file_path = r"G:\Visual_Studio_Work_Place\Python Projects\games_questions.xlsx"

# Load workbook
wb = load_workbook(file_path)
ws = wb.active

# Convert only column D (points) to text
for row in range(2, ws.max_row + 1):
    cell = ws[f"D{row}"]

    if cell.value is not None:
        cell.number_format = "@"
        cell.value = str(cell.value)

# Save file
new_file = r"G:\Visual_Studio_Work_Place\Python Projects\games_questions_fixed.xlsx"
wb.save(new_file)

print("Done! Saved as:", new_file)